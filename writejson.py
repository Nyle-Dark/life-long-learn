import os
import json
from openpyxl import load_workbook


def excel_to_json(excel_path, output_dir):
    """适配当前结构：将单个课程Excel转换为指定格式的JSON"""
    wb = load_workbook(excel_path, read_only=True)
    sheet = wb.active

    result = {
        "课程名称": "",
        "任务模块": "",
        "知识点实体 (Knowledge)": [],
        "技能点实体 (Skill)": [],
        "关联证书 (Certification)": []
    }

    # 提取课程名称：第一行合并单元格是课程标题
    first_cell = sheet.cell(row=1, column=1).value
    if  first_cell and  "课程能力图谱" in first_cell:
        result["课程名称"] = str(first_cell).strip().replace("《", "").replace("》", "").replace("课程能力图谱",
                                                                                               "").strip()
    # 从第三行开始读取数据，跳过表头行
    for row in sheet.iter_rows(min_row=3, values_only=True):
        # 列对应关系：
        # 0:A = 核心工作项目(模块) = 任务模块，同一模块多行复用这个值
        # 3:D = 核心职业能力 = 技能点
        # 4:E = 可考核技能点 = 补充技能点
        # 6:G = 支撑知识点 = 知识点
        core_module, work_task, _, core_ability, check_skill, _, support_know = row[:7]

        # 提取任务模块：当前行有模块名就更新，没有就沿用之前的（合并单元格特性）
        if core_module and str(core_module).strip():
            current_module = str(core_module).strip()
            if not result["任务模块"]:
                result["任务模块"] = current_module

        # 提取所有知识点，按换行分割存入
        if support_know and str(support_know).strip():
            for item in str(support_know).split("\n"):
                item = item.strip()
                # 去除开头的序号
                item_clean = item.split(".", 1)[-1].strip()
                if item_clean:
                    result["知识点实体 (Knowledge)"].append(item_clean)

        # 提取所有技能点，核心职业能力+可考核技能点都存入
        if core_ability and str(core_ability).strip():
            for item in str(core_ability).split("\n"):
                item = item.strip()
                item_clean = item.split(".", 1)[-1].strip()
                if item_clean:
                    result["技能点实体 (Skill)"].append(item_clean)
        if check_skill and str(check_skill).strip():
            for item in str(check_skill).split("\n"):
                item = item.strip()
                item_clean = item.split(".", 1)[-1].strip()
                if item_clean:
                    result["技能点实体 (Skill)"].append(item_clean)

    wb.close()

    # 输出JSON
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=4)
    print(f"转换完成：{output_path}")


def batch_convert(input_dir, output_dir):
    """批量转换课程大纲文件夹下的所有Excel"""
    os.makedirs(output_dir, exist_ok=True)
    for file_name in os.listdir(input_dir):
        if file_name.lower().endswith(('.xlsx', '.xls')) and not file_name.startswith('~$'):
            excel_path = os.path.join(input_dir, file_name)
            excel_to_json(excel_path, output_dir)


if __name__ == "__main__":
    # 修改为你的课程大纲文件夹路径
    INPUT_EXCEL_FOLDER = "D:\课程大纲\云计算技术应用"
    OUTPUT_JSON_FOLDER = "D:\课程大纲\云计算技术应用_json"
    batch_convert(INPUT_EXCEL_FOLDER, OUTPUT_JSON_FOLDER)

